#!/usr/bin/env python3
"""
金融庁の金融商品仲介業者登録一覧（Excel）からデータを取得し、
Supabaseのadvisorsテーブルに同期するスクリプト。

データソース: https://www.fsa.go.jp/menkyo/menkyoj/chuukai.xlsx
"""

import json
import os
import re
import sys
import tempfile
from datetime import datetime

import openpyxl
import requests
from dotenv import load_dotenv

# .env.local から読み込み
env_path = os.path.join(os.path.dirname(__file__), '..', '.env.local')
load_dotenv(env_path)

SUPABASE_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
SERVICE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
FSA_XLSX_URL = "https://www.fsa.go.jp/menkyo/menkyoj/chuukai.xlsx"

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

# 都道府県の抽出パターン
PREFECTURE_PATTERN = re.compile(
    r'(北海道|青森県|岩手県|宮城県|秋田県|山形県|福島県|'
    r'茨城県|栃木県|群馬県|埼玉県|千葉県|東京都|神奈川県|'
    r'新潟県|富山県|石川県|福井県|山梨県|長野県|岐阜県|'
    r'静岡県|愛知県|三重県|滋賀県|京都府|大阪府|兵庫県|'
    r'奈良県|和歌山県|鳥取県|島根県|岡山県|広島県|山口県|'
    r'徳島県|香川県|愛媛県|高知県|福岡県|佐賀県|長崎県|'
    r'熊本県|大分県|宮崎県|鹿児島県|沖縄県)'
)


def download_xlsx():
    """金融庁のExcelファイルをダウンロード"""
    print(f"金融庁データをダウンロード中... {FSA_XLSX_URL}")
    resp = requests.get(FSA_XLSX_URL, timeout=30)
    resp.raise_for_status()
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.write(resp.content)
    tmp.close()
    print(f"ダウンロード完了: {len(resp.content)} bytes")
    return tmp.name


def parse_xlsx(filepath):
    """Excelファイルをパースしてアドバイザーデータのリストを返す"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    advisors = []
    # ヘッダー行を探す（「所管」「登録番号」が含まれる行）
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row and row[0] and '所管' in str(row[0]):
            header_row = row_idx
            break

    if not header_row:
        print("ヘッダー行が見つかりません")
        return []

    print(f"ヘッダー行: {header_row}")

    current_jurisdiction = ""
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        # 所管（A列）- 財務局名が入っている場合は更新
        if row[0] and '財務局' in str(row[0]):
            current_jurisdiction = str(row[0]).split('\n')[0].strip()

        # 業者名（D列）が空なら空行 → スキップ
        name = row[3]
        if not name or not str(name).strip():
            continue

        name = str(name).strip()
        reg_number = str(row[1]).strip() if row[1] else ""
        reg_date_raw = row[2]
        corp_number = str(row[4]).strip() if row[4] else ""
        zipcode = str(row[5]).strip() if row[5] else ""
        address = str(row[6]).strip().replace('\n', ' ') if row[6] else ""
        phone = str(row[7]).strip() if row[7] else ""
        corp_or_individual = str(row[8]).strip() if row[8] else ""
        affiliated = str(row[9]).strip().replace('\n', '、') if row[9] else ""

        # 個人は除外（法人のみ取得）
        if corp_or_individual == '個人':
            continue

        # 登録日のパース
        reg_date = ""
        if reg_date_raw:
            if isinstance(reg_date_raw, datetime):
                reg_date = reg_date_raw.strftime("%Y-%m-%d")
            else:
                try:
                    # Excelシリアル値
                    from datetime import timedelta
                    base = datetime(1899, 12, 30)
                    reg_date = (base + timedelta(days=int(reg_date_raw))).strftime("%Y-%m-%d")
                except (ValueError, TypeError):
                    reg_date = str(reg_date_raw)

        # 都道府県の抽出
        prefecture = ""
        m = PREFECTURE_PATTERN.search(address)
        if m:
            prefecture = m.group(1)

        # 所属先から専門分野タグを生成
        specialties = ["資産運用"]
        if "楽天証券" in affiliated:
            specialties.append("楽天証券")
        if "SBI証券" in affiliated:
            specialties.append("SBI証券")
        if "マネックス証券" in affiliated:
            specialties.append("マネックス証券")
        if "ウェルスナビ" in affiliated:
            specialties.append("ロボアドバイザー")

        # 説明文の生成
        desc_parts = [f"金融商品仲介業者（{reg_number}）。"]
        if affiliated:
            desc_parts.append(f"所属金融商品取引業者等：{affiliated}。")
        if reg_date:
            desc_parts.append(f"登録年月日：{reg_date}。")
        description = " ".join(desc_parts)

        advisors.append({
            "name": name,
            "type": "ifa",
            "description": description[:500],  # 長すぎる場合はカット
            "address": address if address else None,
            "prefecture": prefecture if prefecture else None,
            "website_url": None,
            "specialties": specialties,
        })

    print(f"パース完了: {len(advisors)}社（法人のみ）")
    return advisors


def get_existing_advisors():
    """Supabaseから既存データを取得"""
    r = requests.get(
        f"{SUPABASE_URL}/rest/v1/advisors?select=id,name,type",
        headers=HEADERS
    )
    r.raise_for_status()
    return {a['name']: a for a in r.json()}


def upsert_advisors(advisors):
    """Supabaseにデータをupsert"""
    existing = get_existing_advisors()
    print(f"既存データ: {len(existing)}件")

    new_advisors = []
    skipped = 0

    for a in advisors:
        if a['name'] in existing:
            skipped += 1
            continue
        new_advisors.append(a)

    print(f"新規追加: {len(new_advisors)}件, スキップ（既存）: {skipped}件")

    if not new_advisors:
        print("追加するデータはありません")
        return

    # バッチで投入（50件ずつ）
    batch_size = 50
    total_added = 0
    for i in range(0, len(new_advisors), batch_size):
        batch = new_advisors[i:i + batch_size]
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/advisors",
            headers=HEADERS,
            data=json.dumps(batch)
        )
        if r.status_code in [200, 201]:
            total_added += len(r.json())
            print(f"  バッチ {i // batch_size + 1}: {len(r.json())}件追加")
        else:
            print(f"  バッチ {i // batch_size + 1} エラー: {r.status_code} - {r.text[:200]}")

    print(f"\n完了！ 合計 {total_added}件 追加")


def main():
    filepath = download_xlsx()
    try:
        advisors = parse_xlsx(filepath)
        if advisors:
            upsert_advisors(advisors)
    finally:
        os.unlink(filepath)


if __name__ == "__main__":
    main()
